
from numpy import row_stack
import pandas as pd
import customtkinter as ctk
from tkinter import *
import tkinter as tk
from openpyxl import load_workbook
from CTkListbox import *
from tkinter import ttk



path="E:\College\C++ Codes\Khat3am_Project\\asdasdasdas.xlsx"  # Adjust it as your excel path



wb = load_workbook( path, data_only=True)
ws = wb.active
df = pd.read_excel(path)


calculation = 0
foul = 0

  





def open_new_window():
    ctk.set_appearance_mode("dark")
    new_window = ctk.CTkToplevel(root)
    new_window.title("Final Order")
    new_window.geometry("400x350")
    new_window.grid_columnconfigure((4), weight=2)
    new_window.grid_rowconfigure((4), weight=2)
    total_cost = 10
    foul_total = 0
    tamea_total = 0
    patats_total = 0
    tamea_ma7sheya_total = 0
    patats_sury_total = 0
    kebda_total = 0
    mix_khat3am_total = 0
    foul_sogok_total = 0
    Final_order = ctk.CTkTextbox(new_window, height=120, width=405, font=("Arial", 18))
    Final_order.grid(row=1, column=0, sticky="n", padx=20, pady=0, columnspan=500)
    for row in range(2, 49):  # Adjust the range to match your requirements
        column_J = ws["J" + str(row)].value  # Name Column
        value_A = ws["A" + str(row)].value  # Name Row
        SUM = ws["I"+str(row)].value*5.5+ws["H"+str(row)].value*5.5+ws["G"+str(row)].value*11+ws["F"+str(row)].value*6.5+ws["E"+str(row)].value*25+ws["D"+str(row)].value*18+ws["C"+str(row)].value*13+ws["B"+str(row)].value*13
        total_cost = total_cost + SUM


        ws["A" + str(row)] = SUM         #????????


        foul_total += ws["I" + str(row)].value
        ws["I49"] = foul_total
        tamea_total += ws["H" + str(row)].value
        ws["H49"] = tamea_total
        patats_total += ws["G" + str(row)].value
        ws["G49"] = patats_total
        tamea_ma7sheya_total += ws["F" + str(row)].value
        ws["F49"] = tamea_ma7sheya_total
        patats_sury_total += ws["E" + str(row)].value
        ws["E49"] = patats_sury_total
        kebda_total += ws["D" + str(row)].value
        ws["D49"] = kebda_total
        mix_khat3am_total += ws["C" + str(row)].value
        ws["C49"] = mix_khat3am_total
        foul_sogok_total += ws["B" + str(row)].value
        ws["B49"] = foul_sogok_total
        text_result.delete(1.0, "end")
        text_Tperson.delete(1.0, "end")
        text_total.delete(1.0 , "end")
        order_per_person_box.delete(1.0 , "end")
        hnwz3()
        

        if column_J is not None and value_A is not None and value_A != 0:
            Final_order.insert(1.0, f"{column_J} , Pay : {value_A}\n")
    All_types = ctk.CTkTextbox(new_window, height=100, width=400, font=("Arial", 20))
    All_types.grid(row=0, column=0, padx=20, pady=10, columnspan=5)
    Total = ctk.CTkTextbox(new_window, height=70, width=400, font=("Arial", 20))
    Total.grid(row=2, column=0, padx=20, pady=10, columnspan=5)
    Total.insert(1.0, f"Total : {total_cost}")
    for column in range(66, 74):
        row_51 = "49"
        all_totals = ws[chr(column) + row_51].value
        type_per_num = ws[chr(column) + "1"].value
        if all_totals is not None and all_totals != 0:
            All_types.insert(1.0, f"{type_per_num} : {all_totals}\n")
        
            
    wb.save(path)


names = []
for i in range(2, 49):
    x = ws["J" + str(i)].value
    names.append(x)

def Name_select():
    global Var_row
    text_result.delete(1.0, "end")
    text_Tperson.delete(1.0, "end")
    text_total.delete(1.0 , "end")
    target_name = combo_names.get()
    text_total.delete(1.0 , "end")
    text_total.insert(1.0 , target_name)
    result = df[df["الاسم"] == target_name]
    if not result.empty:
        row_index = result.index[0]
        ROW = row_index + 2
        Var_row = ROW


def clear_field():
    for row in range(2 , 49):
        for column in range(65 , 74):
            ws[str(chr(column)) + str(row)].value = 0
    wb.save(path)


def search_items():
    search_value = variable.get().lower()
    if not search_value.strip():
        combo_names["values"] = item_names
    else:
        value_to_display = [
            str(value) for value in item_names if search_value in str(value).lower()
        ]
        combo_names["values"] = value_to_display


ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("E:\College\C++ Codes\Khat3am_Project\Color_themes\Theme 3.json")
root = ctk.CTk()
root.geometry("800x730")
root.title("tablya")
root.after(201, lambda :root.iconbitmap("E:\College\C++ Codes\Khat3am_Project\photo_2024-01-18_19-18-03.ico"))
root.resizable(False, False)
root.grid_columnconfigure((8), weight=2)
root.grid_rowconfigure((4), weight=2)


foul_name = "  فول   "

frame = ctk.CTkScrollableFrame(root, height=196, width=400)   #ORDER FRAME
frame.grid_columnconfigure((4), weight=2)
frame.grid_rowconfigure((8), weight=2)
frame.grid(row=0, column=4, padx=20, pady=20, sticky="ne", columnspan=10)


frame2 = ctk.CTkFrame(root, width=5000, height=566)           #results frame
frame2.grid_columnconfigure((2), weight=2)
frame2.grid_rowconfigure((2), weight=2)
frame2.grid(row=0, column=0, padx=10, pady=10 ,sticky="WN")


frame3 = ctk.CTkFrame(root, height=244, width=460)            #BUTTON FRAME
frame3.grid_columnconfigure((5), weight=2)
frame3.grid_rowconfigure((10), weight=2)
frame3.grid(row=0, column=0, padx=10, pady=525, sticky="WE")


frame4 = ctk.CTkFrame(root, width=500, height=320 )   #NAME SELECTING
frame4.grid_columnconfigure((4), weight=2)
frame4.grid_rowconfigure((3), weight=3)
frame4.grid(row=0, column=4, padx=10, pady=240, sticky="EN")

text_result = ctk.CTkTextbox(frame2, height=400, width=250, font=("Arial", 30))
text_result.grid(row=2, column=0, padx=10, pady=10, columnspan=10, sticky="nw")
text_Tperson = ctk.CTkTextbox(frame2, height=400, width=100, font=("Arial", 28))
text_Tperson.grid(row=2, column=1, padx=10, pady=10, columnspan=10, sticky="ne")
text_total = ctk.CTkTextbox(frame2, height=68, width=350, font=("Arial", 28))
text_total.grid(row=0, column=0, padx=10, pady=10, columnspan=10)




# من هنا الجديد بتاعي زودت ست زراير


def add_to_calculation(column , type ):
    global main
    global column2
    column2 = column
    main = ws[column + str(Var_row)].value
    text_result.insert(1.0 , type + "\n")
    text_Tperson.insert(1.0 , "\n")


def plus():
    global column2
    global main
    int(main)
    main +=1
    ws[column2 + str(Var_row)].value +=1
    ngm3()
    text_Tperson.delete(1.0 , 2.0)
    text_Tperson.insert(1.0, "\n")
    text_Tperson.insert(1.0 ,main)


def minus():
    global column2
    global main
    int(main)
    main -=1
    ws[column2 + str(Var_row)].value -=1
    ngm3()
    text_Tperson.delete(1.0 , 2.0)
    text_Tperson.insert(1.0, "\n")
    text_Tperson.insert(1.0 , main)
                        
Fbtn1 = ctk.CTkButton(
    frame,
    text=foul_name,
    command=lambda: add_to_calculation("I", " فول "),
    width=60,
    height=40,
    font=("Simplified Arabic", 20),
)
Fbtn1.grid(row=0, column=0, padx=5, pady=5)

Fbtn2 = ctk.CTkButton(
    frame,
    text="طعميه",
    command=lambda: add_to_calculation("H", " طعميه "),
    width=60,
    height=40,
    font=("Simplified Arabic", 20),
)
Fbtn2.grid(row=0, column=1, padx=5, pady=5)

Fbtn3 = ctk.CTkButton(
    frame,
    text="بطاطس",
    command=lambda: add_to_calculation("G", " بطاطس "),
    width=60,
    height=40,
    font=("Simplified Arabic", 20),
)
Fbtn3.grid(row=0, column=2, padx=5, pady=5)

Fbtn4 = ctk.CTkButton(
    frame,
    text="طعميه محشيه",
    command=lambda: add_to_calculation("F", " طعميه محشيه "),
    width=60,
    height=40,
    font=("Simplified Arabic", 20),
)
Fbtn4.grid(row=0, column=3, padx=0, pady=5)


Fbtn5 = ctk.CTkButton(
    frame,
    text="بطاطس سوري",
    command=lambda: add_to_calculation("E", "بطاطس سوري "),
    width=60,
    height=40,
    font=("Simplified Arabic", 20),
)
Fbtn5.grid(row=1, column=3, padx=5, pady=5)


Fbtn6 = ctk.CTkButton(
    frame,
    text="كبده",
    command=lambda: add_to_calculation("D", " كبده "),
    width=60,
    height=40,
    font=("Simplified Arabic", 20),
)
Fbtn6.grid(row=1, column=2, padx=5, pady=5)


Fbtn7 = ctk.CTkButton(
    frame,
    text="ميكس ختعم",
    command=lambda: add_to_calculation("C", " ميكس ختعم "),
    width=60,
    height=40,
    font=("Simplified Arabic", 20),
)
Fbtn7.grid(row=1, column=1, padx=5, pady=5)


Fbtn8 = ctk.CTkButton(
    frame,
    text="فول سجق",
    command=lambda: add_to_calculation("B", " فول سجق"),
    width=60,
    height=40,
    font=("Simplified Arabic", 20),
)
Fbtn8.grid(row=1, column=0, padx=5, pady=5)







btn_clear = ctk.CTkButton(
    frame3, text="Clear", command=lambda: clear_field(), height=80, width=80
)
btn_clear.grid(row=1, column=1, padx=20, pady=20)

btn_next = ctk.CTkButton(
    frame3, height=80, width=80, fg_color="#8B0000" ,hover_color="red",text="self_destruction", command=lambda: next_person()
)
btn_next.grid(row=1, column=0, padx=20, pady=20, sticky="w")


btn_order = ctk.CTkButton(
    frame3, height=80, width=80, command=lambda: open_new_window(), text="Order"
)
btn_order.grid(row=1, column=2, padx=20, pady=20)

"""combo=ctk.CTkComboBox(root , height=50 , width=250 , values=Names , font=("Arial",20))
combo.grid(row=0 , column=10 , padx=10 , pady=10 , sticky="e")
"""
"""plus_food = ctk.CTkButton(root,height=60 , width=60 ,text="+" , command=lambda:fplus_food("+"))
plus_food.grid(row=0 , column=4 , padx=20 , pady=20 , sticky="e" , columnspan=2)
"""


add_btn = ctk.CTkButton(
    frame4, height=50, width=50, text=" + ", font=("Arial", 26), corner_radius=90 , command=lambda: plus()
)
add_btn.grid(row=3, column=1, padx=20, pady=0, columnspan=1)

minus_btn = ctk.CTkButton(
    frame4, height=50, width=50, text=" - ", font=("Arial", 28), corner_radius=90 , command=lambda: minus()
)
minus_btn.grid(row=3, column=0, padx=20, pady=0, columnspan=1)

new_var = combo_names = ctk.CTkComboBox(
    frame4,
    height=20,
    width=300,
    font=("Arial", 16),
    values=names, 
)
    
new_var

combo_names.grid(row=0, column=0, padx=20, pady=0, sticky="e", columnspan=5)


item_names = names
variable = StringVar()
combo_names["values"] = item_names

# Inside your class definition
class YourClassName:
    def __init__(self, root):
        # ... your existing code ...


     variable = StringVar()
entry = ctk.CTkEntry(frame4, textvariable=variable, height=20, width=250)
entry.grid(row=1, column=0, padx=20, pady=0, columnspan=5, sticky="e")

result_listbox = CTkListbox(frame4, height=120, width=300)
result_listbox.grid(row=2, column=0, padx=00, pady=10, columnspan=5)

    # Bind the entry widget to the update_results function
entry.bind("<KeyRelease>", lambda event, entry=entry, result_listbox=result_listbox: update_results(event, entry, result_listbox))
result_listbox.bind("<<ListboxSelect>>", lambda event: handle_selection(result_listbox))
# Add this function outside the class definition
def update_results(event, entry, result_listbox):
    # Get the letter entered by the user
    search_letter = entry.get()

    # Filter the DataFrame to include only rows where the letter is present in the 'الاسم' column
    result_df = df[df['الاسم'].str.contains(search_letter, case=False, na=False)]

    # Display the result in the Listbox widget
    result_listbox.delete(0, ctk.END)
    for name in result_df['الاسم']:
        result_listbox.insert(ctk.END, name)



        


def handle_selection(result_listbox):
        # Get the selected index from the Listbox
        selected_index = result_listbox.curselection()

        if selected_index:
            # Get the name associated with the selected index
            selected_name = result_listbox.get(selected_index)
            global Var_row
            text_result.delete(1.0, "end")
            text_Tperson.delete(1.0, "end")
            text_total.delete(1.0 , "end")
            selected_name = result_listbox.get(selected_index)
            text_total.delete(1.0 , "end")
            text_total.insert(1.0 , selected_name)
            result = df[df["الاسم"] == selected_name]
            if not result.empty:
                    row_index = result.index[0]
                    ROW = row_index + 2
                    Var_row = ROW


Search_btn = ctk.CTkButton(
    frame4,
    text="Select",
    command=lambda: Name_select(),
    height=50,
    width=100,
    corner_radius=20,
)
Search_btn.grid(row=3, column=3, padx=30, pady=0)

Makers = ctk.CTkLabel(frame3 , text="  Ahmed Baalash, Yousef Fady, Ahmed Fouda" , font=("Mistral" , 24) , text_color="#332941")
Makers.grid(row=0 , column=0 , padx=5 , pady=10 , sticky='w' , columnspan=5)

def ngm3():
 for row in range(2 , 49):
        for column in range(65 , 74):
            ws[str(chr(65)) + str(row)].value = ws["I"+str(row)].value*5.5+ws["H"+str(row)].value*5.5+ws["G"+str(row)].value*11+ws["F"+str(row)].value*6.5+ws["E"+str(row)].value*25+ws["D"+str(row)].value*18+ws["C"+str(row)].value*13+ws["B"+str(row)].value*13
wb.save(path)


order_per_person_box= ctk.CTkTextbox(frame4, height=173, width=350, font=("Arial", 18))
order_per_person_box.grid(row=5, column=0, sticky="n", padx=20, pady=10, columnspan=500) 
    

def has_non_zero_quantity(quantities):
    return any(quantity != 0 for quantity in quantities)

def print_non_zero_quantities(name, headers, quantities):
    new_window = tk.Toplevel(root)
    new_window.title("New Window")
    new_window.geometry("500x420")
    Final_order = ctk.CTkTextbox(new_window, height=120, width=350, font=("Arial", 18))
    Final_order.grid(row=1, column=0, sticky="n", padx=20, pady=0, columnspan=500) 
    if has_non_zero_quantity(quantities):
        Final_order.insert(ctk.END, f"Name: {name}\n")
        for header, quantity in zip(headers, quantities):
            if quantity != 0:
                Final_order.insert(ctk.END, f"{header}: {quantity}\n")
        Final_order.insert(ctk.END, "\n")

    # Get headers from the first row
    headers = [cell.value for cell in ws[1]]

    # Iterate through each row starting from the second row
    for row in range(2, 49):
        name = ws.cell(row=row, column=10).value
        quantities = [cell.value for cell in ws[row][2:9]]
        print_non_zero_quantities(name, headers[2:9], quantities)


def next_person():
  print_non_zero_quantities(name, headers, quantities)        
headers = [cell.value for cell in ws[1]]

    # Iterate through each row starting from the second row
for row in range(2, 49):
        name = ws.cell(row=row, column=10).value
        quantities = [cell.value for cell in ws[row][2:9]]




def order_per_person(name, headers, quantities):
  if has_non_zero_quantity(quantities):
        order_per_person_box.insert(tk.END, f"Name: {name}\n")
        for header, quantity in zip(headers, quantities):
            if quantity != 0:
                order_per_person_box.insert(tk.END, f"{header}: {quantity}\n")
        order_per_person_box.insert(tk.END, "\n")
def hnwz3():
     # Get headers from the first row
        headers = [cell.value for cell in ws[1]]

    # Iterate through each row starting from the second row
        for row in range(2, 49):
         name = ws.cell(row=row, column=10).value
         quantities = [cell.value for cell in ws[row][2:9]]
         order_per_person(name, headers[2:9], quantities)

root.mainloop()
